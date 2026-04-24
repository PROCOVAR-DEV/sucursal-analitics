import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

# =========================
# CONFIGURACIÓN
# =========================
# NOMBRE EXACTO DE TU ARCHIVO EXCEL
ARCHIVO_ENTRADA = "Venta Vendedores 01-17 ABRIL.xlsx"

SUPERVISOR_NOMBRE = "Camaguey"
DIAS_LABORABLES = 20
TARGETS = {"HL": 920, "CCC": 500}
CURVA_VENTA = {"S1": 4, "S2": 5, "S3": 6.5, "S4": 6.5}
FRECUENCIA = {"S1": 1, "S2": 1, "S3": 1, "S4": 1}

GESTORES_CONFIG = {
    "ALEXANDER": {"nombre_completo": "Alexander Padron", "agencia": "Camaguey", "sector": "Zona 1", "cuota_hl": 178.5, "cuota_ccc": 100, "incluir_ccc": True},
    "DEYANIRA": {"nombre_completo": "Deyanira Zaldivar", "agencia": "Camaguey", "sector": "Zona 2", "cuota_hl": 178.5, "cuota_ccc": 100, "incluir_ccc": True},
    "GEORLIS": {"nombre_completo": "Georlis Cardenas", "agencia": "Camaguey", "sector": "Zona 3", "cuota_hl": 178.5, "cuota_ccc": 100, "incluir_ccc": True},
    "JEAN MICHEL": {"nombre_completo": "Jean Ramos", "agencia": "Camaguey", "sector": "Zona 4", "cuota_hl": 178.5, "cuota_ccc": 100, "incluir_ccc": True},
    "MAYLEN": {"nombre_completo": "Maylen Remon", "agencia": "Camaguey", "sector": "Zona 6", "cuota_hl": 178.5, "cuota_ccc": 100, "incluir_ccc": True},
}
GESTORES_ORDEN = ["ALEXANDER", "DEYANIRA", "GEORLIS", "JEAN MICHEL", "MAYLEN"]

# =========================
# CARGA DE DATOS (DESDE EL EXCEL)
# =========================

def load_data(archivo):
    if not os.path.exists(archivo):
        raise FileNotFoundError(f"NO SE ENCUENTRA EL ARCHIVO: {archivo}")
    
    print(f"Leyendo archivo: {archivo} ...")
    
    # 1. Leer Supervisor (Buscamos la fila donde empieza la tabla)
    # Leemos sin cabecera primero para encontrar la fila que dice "Gestor"
    df_temp = pd.read_excel(archivo, sheet_name="Supervisor", header=None)
    
    # Buscar en qué fila está la palabra "Gestor" (columna A)
    start_row = 0
    for idx, row in df_temp.iterrows():
        if str(row[0]).strip() == "Gestor":
            start_row = idx
            break
            
    # Volvemos a leer saltando las filas vacías
    supervisor_df = pd.read_excel(archivo, sheet_name="Supervisor", skiprows=start_row)
    
    # 2. Leer hojas de Gestores
    gestores_data = {}
    xls = pd.ExcelFile(archivo)
    for sheet in xls.sheet_names:
        nombre_hoja = sheet.strip().upper() # Normalizar nombre hoja
        if nombre_hoja in GESTORES_CONFIG:
            print(f"  - Leyendo hoja: {sheet}")
            gestores_data[nombre_hoja] = pd.read_excel(archivo, sheet_name=sheet)
            
    return supervisor_df, gestores_data

# Ejecutar carga
try:
    supervisor_df, gestores_data = load_data(ARCHIVO_ENTRADA)
except Exception as e:
    print(f"\nERROR CRÍTICO LEYENDO EL EXCEL:\n{e}")
    exit()

# =========================
# LÓGICA DE CÁLCULO
# =========================

def get_week_number(date_obj):
    if pd.isna(date_obj): return None
    try:
        day = date_obj.day
        return min(((day - 1) // 7) + 1, 5)
    except:
        return None

def calculate_weekly(df):
    w_hl = {f"S{i}": 0.0 for i in range(1, 6)}
    w_ccc = {f"S{i}": set() for i in range(1, 6)}
    
    if df.empty: return w_hl, {k:0 for k in w_hl}

    # Normalizar columnas (quitar espacios extra)
    df.columns = [str(c).strip() for c in df.columns]

    # Buscar columnas clave (insensible a mayúsculas)
    col_date = next((c for c in df.columns if "fecha" in c.lower()), None)
    col_hl = next((c for c in df.columns if "hectolitros" in c.lower()), None)
    col_cli = next((c for c in df.columns if "socio" in c.lower()), None)
    
    if col_date:
        df[col_date] = pd.to_datetime(df[col_date], errors='coerce')
        for _, row in df.iterrows():
            d = row[col_date]
            if pd.notna(d):
                wk_num = get_week_number(d)
                if wk_num:
                    wk = f"S{wk_num}"
                    # Sumar HL
                    if col_hl and wk in w_hl: 
                        val = row.get(col_hl, 0)
                        if pd.notna(val) and isinstance(val, (int, float)):
                            w_hl[wk] += float(val)
                    # Sumar CCC (Set)
                    if col_cli and wk in w_ccc: 
                        cli = row.get(col_cli)
                        if pd.notna(cli): 
                            w_ccc[wk].add(cli)
                        
    return w_hl, {k: len(v) for k,v in w_ccc.items()}

def safe_int(v):
    try: return int(round(float(v),0))
    except: return 0

# =========================
# GENERAR REPORTE EXCEL
# =========================
print("Generando reporte...")

wb = Workbook()
ws = wb.active
ws.title = "Reporte de Ventas"

# Estilos
fill_header = PatternFill("solid", fgColor="1F2937")
fill_total = PatternFill("solid", fgColor="374151") 
fill_yellow = PatternFill("solid", fgColor="FBBF24")
fill_row1 = PatternFill("solid", fgColor="E5E7EB")
fill_row2 = PatternFill("solid", fgColor="F9FAFB")

font_white = Font(color="FFFFFF", bold=True, size=8)
font_black = Font(color="000000", size=8)
font_yellow = Font(color="000000", bold=True, size=8)
font_cyan = Font(color="0891B2", bold=True, size=9)
font_total = Font(color="FFFFFF", bold=True, size=8)

border = Border(left=Side(style='thin', color='9CA3AF'), 
                right=Side(style='thin', color='9CA3AF'), 
                top=Side(style='thin', color='9CA3AF'), 
                bottom=Side(style='thin', color='9CA3AF'))

align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_l = Alignment(horizontal='left', vertical='center')

# Cabecera
ws["A1"] = "Reporte de ventas"; ws["A1"].font = Font(color="0891B2", size=14, bold=True)
ws["A2"] = "Supervisor"; ws["A2"].font = font_cyan
ws["B2"] = SUPERVISOR_NOMBRE; ws["B2"].font = Font(color="DC2626", size=10, bold=True)

row = 4
# --- SECCION HL ---
ws.cell(row, 1, "Target HL").font = font_cyan
ws.cell(row, 2, TARGETS["HL"]).font = font_black
ws.cell(row, 5, DIAS_LABORABLES).font = font_black
ws.cell(row, 7, "Curva Venta").font = font_black
for i, v in enumerate(CURVA_VENTA.values()): ws.cell(row, 8+i*2, v).font = font_black
ws.cell(row, 18, DIAS_LABORABLES).font = font_black

row += 1
headers = ["Indicador", "Agencias", "Vendedor", "Sector", "Ventas\nmes\npasado", "Mix de\nventa", 
           "Cuota\nmes", "Cuota\nS1", "Real\nS1", "Cuota\nS2", "Real\nS2", "Cuota\nS3", "Real\nS3", 
           "Cuota\nS4", "Real\nS4", "Cuota\nS5", "Real\nS5", "Cuota\nal día", "Real\nMes", 
           "Real S3", "", "Real\nMes", ""]

for c, h in enumerate(headers, 1):
    cell = ws.cell(row, c, h)
    cell.fill = fill_header; cell.font = font_white; cell.alignment = align_c; cell.border = border

row += 1
totales_hl = {"cuota":0, "real":0}
totales_semanales_hl = {f"S{i}":{"cuota":0, "real":0} for i in range(1,6)}

for idx, gestor in enumerate(GESTORES_ORDEN):
    cfg = GESTORES_CONFIG[gestor]
    # Usar el nombre de la hoja en mayúsculas
    whl, _ = calculate_weekly(gestores_data.get(gestor, pd.DataFrame()))
    
    # Extraer Real Mes del Supervisor DF
    try:
        # Filtrar de forma segura
        fila_sup = supervisor_df[supervisor_df['Gestor'].astype(str).str.upper() == gestor]
        if not fila_sup.empty:
            real_mes = float(fila_sup['Total Hectolitros'].values[0])
        else:
            real_mes = sum(whl.values())
    except Exception as e:
        print(f"Advertencia: No se pudo leer 'Total Hectolitros' para {gestor} en hoja Supervisor. Calculando manual.")
        real_mes = sum(whl.values())
        
    cuota = cfg["cuota_hl"]
    cuotas_sem = {k: round(cuota * v / sum(CURVA_VENTA.values()),0) for k,v in CURVA_VENTA.items()}
    
    fill = fill_row1 if idx % 2 == 0 else fill_row2
    
    # Datos Fijos
    vals = ["HL", cfg["agencia"], cfg["nombre_completo"], cfg["sector"], 0, "0%", safe_int(cuota)]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill; cell.font = font_black; cell.border = border
        cell.alignment = align_l if c in [2,3,4] else align_c

    col = 8
    # Semanas
    for i in range(1,6):
        k = f"S{i}"
        c_sem = safe_int(cuotas_sem.get(k, 0))
        r_sem = safe_int(whl.get(k, 0))
        
        ws.cell(row, col, c_sem).fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
        ws.cell(row, col, r_sem).fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
        
        totales_semanales_hl[k]["cuota"] += c_sem
        totales_semanales_hl[k]["real"] += r_sem
        
    # Totales Fila
    ws.cell(row, col, safe_int(cuota)).fill=fill_yellow; ws.cell(row, col).font=font_yellow; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, safe_int(real_mes)).fill=fill_yellow; ws.cell(row, col).font=font_yellow; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    
    # Indicadores
    pct_s3 = (whl["S3"]/cuotas_sem.get("S3",1)*100) if cuotas_sem.get("S3",0)>0 else 0
    color_s3 = "00FF00" if pct_s3>=100 else "FFFF00" if pct_s3>=80 else "FF0000"
    ws.cell(row, col, f"{pct_s3:.0f}%").fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, "●").fill=fill; ws.cell(row, col).font=Font(color=color_s3, size=10, bold=True); ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1

    pct_mes = (real_mes/cuota*100) if cuota>0 else 0
    color_mes = "00FF00" if pct_mes>=100 else "FFFF00" if pct_mes>=80 else "FF0000"
    ws.cell(row, col, f"{pct_mes:.0f}%").fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, "●").fill=fill; ws.cell(row, col).font=Font(color=color_mes, size=10, bold=True); ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1

    totales_hl["cuota"]+=cuota
    totales_hl["real"]+=real_mes
    row += 1

# Total HL
for c in range(1, 25): ws.cell(row, c).fill = fill_total; ws.cell(row, c).border = border
ws.cell(row, 1, "Total").font = font_total
ws.cell(row, 7, safe_int(totales_hl["cuota"])).font = font_total; ws.cell(row, 7).alignment = align_c

col = 8
for i in range(1,6):
    ws.cell(row, col, safe_int(totales_semanales_hl[f"S{i}"]["cuota"])).font=font_total; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, safe_int(totales_semanales_hl[f"S{i}"]["real"])).font=font_total; ws.cell(row, col).alignment=align_c; col+=1

ws.cell(row, 18, safe_int(totales_hl["cuota"])).fill=fill_yellow; ws.cell(row, 18).font=font_yellow; ws.cell(row, 18).alignment=align_c
ws.cell(row, 19, safe_int(totales_hl["real"])).fill=fill_yellow; ws.cell(row, 19).font=font_yellow; ws.cell(row, 19).alignment=align_c

row += 3
# --- SECCION CCC ---
ws.cell(row, 1, "Target CCC").font = font_cyan
ws.cell(row, 2, TARGETS["CCC"]).font = font_black
ws.cell(row, 5, 5).font = font_black
ws.cell(row, 7, "Frecuencia").font = font_black
for i, v in enumerate(FRECUENCIA.values()): ws.cell(row, 8+i*2, v).font = font_black
ws.cell(row, 18, 5).font = font_black

row += 1
for c, h in enumerate(headers, 1):
    cell = ws.cell(row, c, h)
    cell.fill = fill_header; cell.font = font_white; cell.alignment = align_c; cell.border = border

row += 1
totales_ccc = {"cuota":0, "real":0}
totales_semanales_ccc = {f"S{i}":{"cuota":0, "real":0} for i in range(1,6)}

for idx, gestor in enumerate(GESTORES_ORDEN):
    cfg = GESTORES_CONFIG[gestor]
    if not cfg["incluir_ccc"]: continue
    
    _, wccc = calculate_weekly(gestores_data.get(gestor, pd.DataFrame()))
    real_ccc = sum(wccc.values())
    cuota = cfg["cuota_ccc"]
    cuotas_sem = {k: round(cuota * v / sum(FRECUENCIA.values()),0) for k,v in FRECUENCIA.items()}
    
    fill = fill_row1 if idx % 2 == 0 else fill_row2
    
    vals = ["CCC", cfg["agencia"], cfg["nombre_completo"], cfg["sector"], 0, "0%", safe_int(cuota)]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row, c, v)
        cell.fill = fill; cell.font = font_black; cell.border = border
        cell.alignment = align_l if c in [2,3,4] else align_c

    col = 8
    for i in range(1,6):
        k = f"S{i}"
        c_sem = safe_int(cuotas_sem.get(k, 0))
        r_sem = safe_int(wccc.get(k, 0))
        
        ws.cell(row, col, c_sem).fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
        ws.cell(row, col, r_sem).fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
        
        totales_semanales_ccc[k]["cuota"] += c_sem
        totales_semanales_ccc[k]["real"] += r_sem
        
    ws.cell(row, col, safe_int(sum(cuotas_sem.values()))).fill=fill_yellow; ws.cell(row, col).font=font_yellow; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, safe_int(real_ccc)).fill=fill_yellow; ws.cell(row, col).font=font_yellow; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    
    # Indicadores
    pct_s3 = (wccc["S3"]/cuotas_sem.get("S3",1)*100) if cuotas_sem.get("S3",0)>0 else 0
    color_s3 = "00FF00" if pct_s3>=100 else "FFFF00" if pct_s3>=80 else "FF0000"
    ws.cell(row, col, f"{pct_s3:.0f}%").fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, "●").fill=fill; ws.cell(row, col).font=Font(color=color_s3, size=10, bold=True); ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1

    pct_mes = (real_ccc/cuota*100) if cuota>0 else 0
    color_mes = "00FF00" if pct_mes>=100 else "FFFF00" if pct_mes>=80 else "FF0000"
    ws.cell(row, col, f"{pct_mes:.0f}%").fill=fill; ws.cell(row, col).font=font_black; ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, "●").fill=fill; ws.cell(row, col).font=Font(color=color_mes, size=10, bold=True); ws.cell(row, col).border=border; ws.cell(row, col).alignment=align_c; col+=1

    totales_ccc["cuota"]+=cuota
    totales_ccc["real"]+=real_ccc
    row += 1

# Total CCC
for c in range(1, 25): ws.cell(row, c).fill = fill_total; ws.cell(row, c).border = border
ws.cell(row, 1, "Total").font = font_total
ws.cell(row, 7, safe_int(totales_ccc["cuota"])).font = font_total; ws.cell(row, 7).alignment = align_c

col = 8
for i in range(1,6):
    ws.cell(row, col, safe_int(totales_semanales_ccc[f"S{i}"]["cuota"])).font=font_total; ws.cell(row, col).alignment=align_c; col+=1
    ws.cell(row, col, safe_int(totales_semanales_ccc[f"S{i}"]["real"])).font=font_total; ws.cell(row, col).alignment=align_c; col+=1

ws.cell(row, 18, safe_int(totales_ccc["cuota"])).fill=fill_yellow; ws.cell(row, 18).font=font_yellow; ws.cell(row, 18).alignment=align_c
ws.cell(row, 19, safe_int(totales_ccc["real"])).fill=fill_yellow; ws.cell(row, 19).font=font_yellow; ws.cell(row, 19).alignment=align_c

wb.save("Reporte_Ventas_Final.xlsx")
print("¡Reporte generado correctamente: Reporte_Ventas_Final.xlsx!")